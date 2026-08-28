# 25010 Characteristic: Functional Suitability, Maintainability, Performance Efficiency

import sys
import os
import subprocess

# --- WINDOWS OS CRASH FIX ---
# Prevents OSError: [Errno 22] Invalid argument when Pyinstaller runs windowless
if sys.platform == "win32":
    if sys.stdin is None:
        sys.stdin = open(os.devnull, "r")
    if sys.stdout is None:
        sys.stdout = open(os.devnull, "w")
    if sys.stderr is None:
        sys.stderr = open(os.devnull, "w")

import traceback
import re
import asyncio
from datetime import datetime, timedelta
import zipfile
import threading
import multiprocessing
import io
import shutil
import calendar
import uuid
import platform
import time
from typing import List, Optional
from contextlib import asynccontextmanager

from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Form, Header
from fastapi.security import OAuth2PasswordBearer
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
import starlette.formparsers
import uvicorn

from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy import func, extract, create_engine, inspect

from jose import JWTError, jwt
from passlib.context import CryptContext
import passlib.handlers.bcrypt
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# --- CORE SETTINGS & LOGGING ---
from database import BASE_DIR, SessionLocal, User, FranchiseRecord, AuditLog, SystemSettings, RouteData, get_pht_now
from ml_engine import run_kmeans_clustering, invalidate_ml_cache
from doc_generator import generate_certificate
from extractor import extract_docx_data
from sync_engine import start_lan_sync, get_local_ip, PEERS, CLUSTER_SECRET
from pydantic import BaseModel

crash_log_path = os.path.join(BASE_DIR, "PASADA_CRASH_LOG.txt")

def force_log(msg):
    try:
        with open(crash_log_path, "a", encoding="utf-8") as f:
            f.write(f"{msg}\n")
    except OSError:
        pass

def exception_hook(exc_type, exc_value, exc_traceback):
    err_msg = "".join(traceback.format_exception(exc_type, exc_value, exc_traceback))
    force_log(f"\n[{datetime.now()}] !!! OS-LEVEL FATAL CRASH !!!\n{err_msg}")
    sys.exit(1)

sys.excepthook = exception_hook

force_log(f"\n[{datetime.now()}] --- BOOTING PASADA BACKEND ---")
force_log(f"Running from: {sys.executable}")

if __name__ == "__main__":
    multiprocessing.freeze_support()

force_log("All libraries loaded successfully!")

# --- APP INITIALIZATION ---
starlette.formparsers.MultiPartParser.max_files = 10000
starlette.formparsers.MultiPartParser.max_fields = 10000

def automated_background_tasks():
    while True:
        try:
            now = get_pht_now()
            
            export_dir = os.path.join(BASE_DIR, "exports")
            if os.path.exists(export_dir):
                current_timestamp = time.time()
                for file_name in os.listdir(export_dir):
                    file_path = os.path.join(export_dir, file_name)
                    if os.path.getmtime(file_path) < current_timestamp - 7200:
                        try:
                            os.remove(file_path)
                        except OSError:
                            pass

            track_file = os.path.join(BASE_DIR, "last_backup.txt")
            last_backup = ""
            if os.path.exists(track_file):
                with open(track_file, "r") as f: 
                    last_backup = f.read().strip()
            
            today_str = now.strftime('%Y-%m-%d')
            
            if (now.hour > 16 or (now.hour == 16 and now.minute >= 45)) and last_backup != today_str:
                backup_dir = os.path.join(BASE_DIR, "backups")
                if not os.path.exists(backup_dir):
                    os.makedirs(backup_dir)
                
                db_path = os.path.join(BASE_DIR, 'pasada_production.db')
                if os.path.exists(db_path):
                    zip_name = os.path.join(backup_dir, f"PASADA_Backup_{today_str}.zip")
                    with zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        zipf.write(db_path, os.path.basename(db_path))
                    force_log(f"Automated backup secured: {zip_name}")
                
                with open(track_file, "w") as f: 
                    f.write(today_str)
                    
        except Exception as e:
            force_log(f"Background task engine error: {e}")
            
        time.sleep(60)

@asynccontextmanager
async def lifespan(app: FastAPI):
    threading.Thread(target=automated_background_tasks, daemon=True).start()
    def safe_sync():
        try:
            start_lan_sync()
        except Exception as e:
            force_log(f"LAN Sync initialization deferred: {e}")
    threading.Thread(target=safe_sync, daemon=True).start()
    yield

app = FastAPI(title="PASADA Registry API", lifespan=lifespan)

app.add_middleware(  
    CORSMiddleware,  
    allow_origins=["*"],  
    allow_credentials=False,  
    allow_methods=["*"],  
    allow_headers=["*"],  
    expose_headers=["Content-Disposition"]  
)  
  
from starlette.requests import Request as _StarletteRequest  
from starlette.responses import Response as _StarletteResponse  
  
@app.middleware("http")  
async def allow_private_network_access(request: _StarletteRequest, call_next):  
    if request.method == "OPTIONS":  
        response = _StarletteResponse(status_code=200)  
    else:  
        response = await call_next(request)  
  
    origin = request.headers.get("origin", "*")  
    response.headers["Access-Control-Allow-Origin"] = origin  
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"  
    response.headers["Access-Control-Allow-Headers"] = "*"  
    response.headers["Access-Control-Allow-Private-Network"] = "true"  
    return response

# --- JWT AUTHENTICATION CONFIGURATION ---
SECRET_KEY_PATH = os.path.join(BASE_DIR, "jwt_secret.key")
if os.path.exists(SECRET_KEY_PATH):
    with open(SECRET_KEY_PATH, "r") as f:
        JWT_SECRET_KEY = f.read().strip()
else:
    JWT_SECRET_KEY = str(uuid.uuid4())
    with open(SECRET_KEY_PATH, "w") as f:
        f.write(JWT_SECRET_KEY)

ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_DAYS = 30 

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(days=ACCESS_TOKEN_EXPIRE_DAYS)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, JWT_SECRET_KEY, algorithm=ALGORITHM)

# --- DATABASE INJECTION ---
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def init_settings(db: Session):
    settings = db.query(SystemSettings).first()
    if not settings:
        settings = SystemSettings()
        db.add(settings)
        db.commit()
    return settings

# --- PYDANTIC MODELS ---
class FranchiseCreate(BaseModel):
    sbn_no: Optional[str] = ""
    operator_name: str
    address: str
    motor_no: str
    chassis_no: str
    make: str
    plate_no: str
    route: str
    driving_route: str = ""
    issue_date: Optional[str] = ""
    valid_until: Optional[str] = ""

class UserCreate(BaseModel):
    first_name: str
    last_name: str
    username: str
    password: str
    role: str

class LoginRequest(BaseModel):
    username: str
    password: str

class SettingsUpdate(BaseModel):
    committee_chair: str
    enable_esignature: Optional[bool] = False 

class PasswordUpdate(BaseModel):
    new_password: str

class UsernameUpdate(BaseModel):
    new_username: str

class RouteDataUpdate(BaseModel):
    population: int
    road_length_km: float

class BulkDeleteRequest(BaseModel):
    sbn_list: List[str]

# --- SECURITY DEPENDENCIES ---
def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    try:
        payload = jwt.decode(token, JWT_SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(status_code=401, detail="Invalid token payload")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
        
    user = db.query(User).filter(User.username == username).first()
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user

# --- CENTRALIZED BUSINESS LOGIC HELPERS ---

def get_full_name(user: User) -> str:
    first = getattr(user, 'first_name', "") or ""
    last = getattr(user, 'last_name', "") or ""
    name = f"{first} {last}".strip()
    
    if not name or name == "None" or name == "None None":
        return getattr(user, 'username', "SYSTEM ADMIN")
    return name

def compute_record_status(record: FranchiseRecord) -> str:
    is_vacant = not record.operator_name or str(record.operator_name).strip() == ""
    if is_vacant:
        return "VACANT"
    issue_year = record.issue_date.year if record.issue_date else 0
    current_year = get_pht_now().year
    
    if getattr(record, 'is_active', True) is False or issue_year <= current_year - 2:
        return "REVOKED"
    if issue_year == current_year - 1:
        return "FLAGGED"
    return "ACTIVE"

def attach_status(record: FranchiseRecord):
    record.status = compute_record_status(record)
    return record

def log_action(db: Session, clerk_name: str, action: str, target_id: str, target_route: str, details: str):
    new_log = AuditLog(clerk_name=clerk_name, action=action, target_id=str(target_id), target_route=target_route, details=details)
    db.add(new_log)
    db.commit()

def determine_status(issue_date):
    if not issue_date: return False
    current_year = get_pht_now().year
    return False if issue_date.year <= current_year - 2 else True

def sanitize_plate(plate: str, chassis: str, motor: str) -> str:
    p, c, m = str(plate).strip().upper(), str(chassis).strip().upper(), str(motor).strip().upper()
    invalid_markers = ["NAN", "NONE", "N/A", "NO PLATE", "TBA", "FOR REG", "NEW", "UNREGISTERED", "CHASSIS", "MOTOR", "UNKNOWN"]
    if not p or any(marker in p for marker in invalid_markers):
        return ""
    if p == c or p == m:
        return ""
    return p

def get_fuzzy_col_dict(row_dict, target):
    for col, val in row_dict.items():
        if target in str(col).upper():
            if val is None or str(val).lower() in ["nan", "none", "nat"]:
                return ""
            return str(val).strip()
    return ""

def parse_safe_date(raw_date, fallback_year):
    if raw_date is None or pd.isna(raw_date) or str(raw_date).strip().lower() in ["nan", "nat", "none", ""]:
        return None
    try:
        if isinstance(raw_date, datetime):
            return raw_date
        dt = pd.to_datetime(str(raw_date), errors='coerce', dayfirst=True)
        if pd.isna(dt):
            return None
        return dt.to_pydatetime()
    except (ValueError, TypeError, OverflowError) as e:
        force_log(f"Date parsing failed: {e}")
        return None

def get_base_sbn(sbn):
    sbn = str(sbn).strip().upper()
    # FIX: Restrict year matching to 2-digits OR plausible 4-digit years (19xx, 20xx) so we don't accidentally match the 4-digit sequence "1000"
    match = re.match(r'^([A-Z0-9]+[\-\_]\d+)[\-\_](\d{2}|19\d{2}|20\d{2})$', sbn)
    if match:
        return match.group(1)
    return sbn

def normalize_base_sbn(sbn):
    if pd.isna(sbn) or not sbn: return ""
    sbn = str(sbn).strip().upper()
    # FIX: Restrict year matching here too
    match = re.match(r'^(.*?)[\s\-\_]+(\d{2}|19\d{2}|20\d{2})$', sbn)
    base = match.group(1) if match else sbn
        
    match = re.match(r'^(.*?)[\s\-\_]+(\d+)$', base)
    if match:
        prefix = re.sub(r'[^A-Z0-9]', '', match.group(1))
        num_str = match.group(2)
        padding = 4 if int(num_str) >= 1000 else 3
        return f"{prefix}-{int(num_str):0{padding}d}"

    clean_base = re.sub(r'[^A-Z0-9]', '', base)
    match = re.match(r'^([A-Z]+)(\d+)$', clean_base)
    if match:
        prefix = match.group(1)
        num_str = match.group(2)
        padding = 4 if int(num_str) >= 1000 else 3
        return f"{prefix}-{int(num_str):0{padding}d}"

    return base.replace(" ", "")

def format_sbn_with_year(sbn, issue_date, is_vacant=False):
    base_sbn = normalize_base_sbn(sbn)
    if is_vacant or not issue_date:
        return base_sbn
    try:
        if isinstance(issue_date, str):
            issue_date = datetime.fromisoformat(issue_date.split('T')[0])
        year_suffix = str(issue_date.year)[-2:]
        return f"{base_sbn}-{year_suffix}"
    except Exception:
        return base_sbn

def get_record_year(dt):
    if dt and isinstance(dt, datetime):
        return dt.year
    elif dt and hasattr(dt, 'year'):
        return dt.year
    return 0

def set_if_blank(curr_val, new_val):
    if (not curr_val or not str(curr_val).strip()) and new_val:
        return str(new_val).strip().upper()
    return curr_val

def extract_sbn_integer(sbn):
    base_sbn = get_base_sbn(sbn)
    nums = re.findall(r'\d+', base_sbn)
    if nums:
        return int(nums[-1])
    return None 
    
def get_sbn_sort_key(record):
    val = extract_sbn_integer(record.sbn_no)
    return val if val is not None else 999999

def generate_next_sbn(db: Session, route: str) -> str:
    route_upper = route.strip().upper()
    records = db.query(FranchiseRecord.sbn_no).filter(
        FranchiseRecord.route == route_upper,
        FranchiseRecord.is_deleted == False
    ).all()
    max_num = 0
    prefix = ""
    padding_len = 3
    
    for (sbn_val,) in records:
        if sbn_val:
            sbn_str = str(sbn_val).strip().upper()
            base = get_base_sbn(sbn_str)
            match = re.match(r'^([A-Z0-9]+)[\-\_]?(\d+)', base)
            if match:
                if not prefix: 
                    prefix = match.group(1) 
                num_str = match.group(2)
                padding_len = max(padding_len, len(num_str))
            
            nums = re.findall(r'\d+', base)
            if nums:
                val = int(nums[-1])
                if val > max_num:
                    max_num = val
                    
    if not prefix:
        clean_route = route_upper.replace("TODA", "").strip()
        prefix = clean_route if len(clean_route) <= 4 else clean_route[:3]
        
    next_num = max(max_num + 1, 1)
    # FIX: Safely ensure we don't accidentally skip valid sequences like 1000
    while next_num == 0:
        next_num += 1
        
    return f"{prefix}-{next_num:0{padding_len}d}"

def get_dominant_driving_route(db: Session, route_name: str) -> str:
    records = db.query(FranchiseRecord.driving_route).filter(
        FranchiseRecord.route == route_name.upper(),
        FranchiseRecord.driving_route != None,
        FranchiseRecord.driving_route != "",
        FranchiseRecord.is_deleted == False
    ).all()
    
    valid_routes = [str(r[0]).strip().upper() for r in records if r[0] and str(r[0]).strip()]
    if not valid_routes:
        return ""
        
    from collections import Counter
    most_common = Counter(valid_routes).most_common(1)
    if most_common:
        return most_common[0][0]
    return ""

def deduplicate_records_by_base_sbn(records):
    unique_map = {}
    for r in records:
        if getattr(r, 'is_deleted', False): continue
        base_id = normalize_base_sbn(r.sbn_no)
        r.sbn_no = base_id
        if base_id not in unique_map:
            unique_map[base_id] = r
        else:
            curr_existing = unique_map[base_id]
            curr_year = get_record_year(curr_existing.issue_date)
            new_year = get_record_year(r.issue_date)
            if new_year > curr_year:
                unique_map[base_id] = r
            elif new_year == curr_year and (r.operator_name and str(r.operator_name).strip() and not str(curr_existing.operator_name).strip()):
                unique_map[base_id] = r
    return list(unique_map.values())

def get_all_deduplicated_records(db: Session):
    all_records = db.query(FranchiseRecord).filter(FranchiseRecord.is_deleted == False).all()
    unique_map = {}
    for r in all_records:
        base_id = normalize_base_sbn(r.sbn_no)
        r.sbn_no = base_id
        route_key = (r.route, base_id)
        if route_key not in unique_map:
            unique_map[route_key] = r
        else:
            curr_existing = unique_map[route_key]
            curr_year = get_record_year(curr_existing.issue_date)
            new_year = get_record_year(r.issue_date)
            if new_year > curr_year:
                unique_map[route_key] = r
            elif new_year == curr_year and (r.operator_name and str(r.operator_name).strip() and not str(curr_existing.operator_name).strip()):
                unique_map[route_key] = r
    return list(unique_map.values())

def escape_excel(val):
    if not val:
        return ""
    s = str(val).strip()
    if s.startswith(('=', '+', '-', '@', '\t', '\r')):
        return f"'{s}"
    return s

# --- API ENDPOINTS ---

@app.get("/health")
def health_check():
    return {"status": "online"}

@app.delete("/api/operators/{sbn_no}")
def delete_single_operator(sbn_no: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    clean_sbn = normalize_base_sbn(sbn_no)
    records = db.query(FranchiseRecord).filter(
        ((FranchiseRecord.sbn_no == clean_sbn) | (FranchiseRecord.sbn_no == sbn_no.upper())) &
        (FranchiseRecord.is_deleted == False)
    ).all()
    
    if not records:
        raise HTTPException(status_code=404, detail="Operator record not found.")
        
    count = len(records)
    target_route = None
    for r in records:
        target_route = r.route
        r.is_deleted = True 
        r.updated_at = get_pht_now()
        
    db.commit()
    invalidate_ml_cache(target_route)
    log_action(db, get_full_name(current_user), "DELETE_RECORD", clean_sbn, "ALL", f"Soft deleted operator record {clean_sbn}.")
    return {"message": f"{count} operator(s) deleted successfully."}

@app.post("/api/operators/bulk-delete")
def delete_bulk_operators(payload: BulkDeleteRequest, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    deleted_count = 0
    for sbn_no in payload.sbn_list:
        clean_sbn = normalize_base_sbn(sbn_no)
        records = db.query(FranchiseRecord).filter(
            ((FranchiseRecord.sbn_no == clean_sbn) | (FranchiseRecord.sbn_no == sbn_no.upper())) &
            (FranchiseRecord.is_deleted == False)
        ).all()
        for r in records:
            r.is_deleted = True 
            r.updated_at = get_pht_now()
            deleted_count += 1
    db.commit()
    invalidate_ml_cache() 
    log_action(db, get_full_name(current_user), "BULK_DELETE", "0", "ALL", f"Bulk deleted {deleted_count} operator record(s).")
    return {"message": f"{deleted_count} operator(s) deleted successfully."}

@app.delete("/api/routes/{route_name}")
def delete_entire_route(route_name: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    route_to_delete = route_name.strip().upper()
    records = db.query(FranchiseRecord).filter(FranchiseRecord.route == route_to_delete).all()
    
    if not records:
        raise HTTPException(status_code=404, detail="Route not found or already empty.")
        
    count = 0
    for r in records:
        if not r.is_deleted:
            r.is_deleted = True
            r.updated_at = get_pht_now()
            count += 1
    db.commit()
    invalidate_ml_cache(route_to_delete)
    log_action(db, get_full_name(current_user), "DELETE_ROUTE", "0", route_to_delete, f"Soft deleted entire route line '{route_to_delete}' containing {count} record(s).")
    return {"message": f"Route '{route_to_delete}' deleted successfully."}

@app.post("/admin/refresh-db")
def refresh_database(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    all_records = db.query(FranchiseRecord).filter(FranchiseRecord.is_deleted == False).all()
    route_sbn_map = {}
    deleted_count = 0
    updated_count = 0
    current_year = get_pht_now().year
    
    for r in all_records:
        sbn_str = str(r.sbn_no).strip().upper()
        # FIX: Only drop true placeholders (where the extracted sequence integer is exactly 0)
        sbn_int = extract_sbn_integer(sbn_str)
        if not sbn_str or sbn_int == 0:
            r.is_deleted = True
            r.updated_at = get_pht_now()
            deleted_count += 1
            continue
        
        clean_sbn = normalize_base_sbn(r.sbn_no)
        if r.sbn_no != clean_sbn:
            r.sbn_no = clean_sbn
            updated_count += 1

        if r.issue_date and r.issue_date.month == 1 and r.issue_date.day == 1 and r.issue_date.year == current_year:
            r.issue_date = None
            r.valid_until = None
            r.is_active = False
            updated_count += 1
        
        is_vacant = not r.operator_name or str(r.operator_name).strip() == ""
        if is_vacant:
            if r.is_active != False or r.issue_date is not None:
                r.is_active = False
                r.issue_date = None
                r.valid_until = None
                updated_count += 1
        
        key = (r.route, clean_sbn)
        if key not in route_sbn_map:
            route_sbn_map[key] = r
        else:
            existing = route_sbn_map[key]
            existing_year = get_record_year(existing.issue_date)
            current_year_val = get_record_year(r.issue_date)
            
            if current_year_val > existing_year:
                existing.is_deleted = True
                existing.updated_at = get_pht_now()
                route_sbn_map[key] = r
                deleted_count += 1
            elif current_year_val == existing_year and (r.operator_name and str(r.operator_name).strip() and not str(existing.operator_name).strip()):
                existing.is_deleted = True
                existing.updated_at = get_pht_now()
                route_sbn_map[key] = r
                deleted_count += 1
            else:
                r.is_deleted = True
                r.updated_at = get_pht_now()
                deleted_count += 1
    
    for key, record in route_sbn_map.items():
        is_vacant = not record.operator_name or str(record.operator_name).strip() == ""
        record.sbn_no = format_sbn_with_year(record.sbn_no, record.issue_date, is_vacant)

    unique_routes = {r.route for r in route_sbn_map.values()}
    for r_name in unique_routes:
        dom_route = get_dominant_driving_route(db, r_name)
        if dom_route:
            for r in route_sbn_map.values():
                if r.route == r_name:
                    curr_dr = str(r.driving_route).strip().upper() if r.driving_route else ""
                    if curr_dr == "" or (curr_dr == "POBLACION" and dom_route != "POBLACION"):
                        r.driving_route = dom_route

    db.commit()
    invalidate_ml_cache() 
    log_action(db, get_full_name(current_user), "REFRESH_DB", "0", "ALL", f"Cleaned SBNs, fixed route strings, reversed phantom dates, and removed {deleted_count} duplicates.")
    return {
        "status": "success",
        "message": f"Database refreshed successfully. Deleted {deleted_count} duplicate records. Total remaining across all routes: {len(route_sbn_map)}."
    }

@app.post("/signup")
def signup(user: UserCreate, db: Session = Depends(get_db)):
    if db.query(User).filter(User.username == user.username).first():
        raise HTTPException(status_code=400, detail="Username taken")
    new_user = User(
        first_name=user.first_name, 
        last_name=user.last_name, 
        username=user.username, 
        password_hash=pwd_context.hash(user.password), 
        role=user.role
    )
    db.add(new_user)
    db.commit()
    log_action(db, "SYSTEM ADMIN", "USER_REGISTRATION", "0", "SYSTEM", f"Registered new account for {user.username}")
    return {"message": "Account created"}

@app.post("/token")
def login(payload: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == payload.username).first()
    if not user or not pwd_context.verify(payload.password, user.password_hash):
        raise HTTPException(status_code=400, detail="Incorrect credentials")
    
    full_name = get_full_name(user)
    access_token = create_access_token(data={"sub": user.username})
    
    log_action(db, full_name, "LOGIN", "0", "SYSTEM", "Successful authentication")
    return {
        "access_token": access_token, 
        "token_type": "bearer", 
        "full_name": full_name, 
        "role": user.role
    }

@app.put("/users/password")
def update_password(payload: PasswordUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if len(payload.new_password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")
    current_user.password_hash = pwd_context.hash(payload.new_password)
    db.commit()
    log_action(db, get_full_name(current_user), "UPDATE_SECURITY", "0", "SYSTEM", "Updated account password")
    return {"status": "success"}

@app.put("/users/username")
def update_username(payload: UsernameUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if db.query(User).filter(User.username == payload.new_username.upper()).first():
        raise HTTPException(status_code=400, detail="Username already exists")
    
    old_username = current_user.username
    current_user.username = payload.new_username.upper()
    db.commit()
    
    log_action(db, get_full_name(current_user), "UPDATE_USERNAME", "0", "SYSTEM", f"Changed username from {old_username} to {payload.new_username.upper()}")
    return {"status": "success"}

@app.get("/settings")
def get_settings(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return init_settings(db)

@app.put("/settings")
def update_settings(settings: SettingsUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    sys_settings = init_settings(db)
    sys_settings.committee_chair = settings.committee_chair
    db.commit()
    log_action(db, get_full_name(current_user), "UPDATE_SETTINGS", "0", "SYSTEM", "Updated Global Committee Settings")
    return {"status": "success"}

@app.post("/route_data/{route_name}")
def update_route_data(route_name: str, payload: RouteDataUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    route_info = db.query(RouteData).filter(RouteData.route_name == route_name.upper()).first()
    if not route_info:
        route_info = RouteData(route_name=route_name.upper(), population=payload.population, road_length_km=payload.road_length_km)
        db.add(route_info)
    else:
        route_info.population = payload.population
        route_info.road_length_km = payload.road_length_km
    db.commit()
    invalidate_ml_cache(route_name) 
    log_action(db, get_full_name(current_user), "UPDATE_ROUTE_DATA", "0", route_name.upper(), f"Updated X2 and X3 factors.")
    return {"status": "success", "message": f"Updated demographic data for {route_name.upper()}"}

@app.get("/system/network")
def get_network_status():
    return {"local_ip": get_local_ip(), "connected_peers": list(PEERS)}

@app.get("/api/sync/pull")
def sync_pull(since: str, x_cluster_secret: str = Header(None), db: Session = Depends(get_db)):
    if x_cluster_secret != CLUSTER_SECRET:
        raise HTTPException(status_code=403, detail="Unauthorized cluster request")
        
    target_time = datetime.fromisoformat(since)
    records = db.query(FranchiseRecord).filter(FranchiseRecord.updated_at > target_time).all()
    users = db.query(User).all()
    
    safe_users = []
    for u in users:
        u_dict = {c.name: getattr(u, c.name) for c in u.__table__.columns if c.name != "password_hash"}
        safe_users.append(u_dict)
        
    return {"records": [attach_status(r) for r in records], "users": safe_users}

@app.post("/api/operators")
def create_operator_api(record: FranchiseCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return create_franchise(record, current_user, db)

@app.post("/franchise/")
def create_franchise(record: FranchiseCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    run_kmeans_clustering(db, record.route.upper())

    full_name = get_full_name(current_user)
    current_time = get_pht_now()
    
    if not record.sbn_no or str(record.sbn_no).strip() == "":
        raw_next = generate_next_sbn(db, record.route)
    else:
        raw_next = record.sbn_no

    is_vacant = not record.operator_name or str(record.operator_name).strip() == ""
    
    manual_issue = parse_safe_date(record.issue_date, current_time.year) if record.issue_date else None
    manual_valid = parse_safe_date(record.valid_until, current_time.year) if record.valid_until else None

    actual_issue_date = manual_issue if manual_issue else (None if is_vacant else current_time)
    actual_valid_until = manual_valid if manual_valid else (None if is_vacant else datetime(current_time.year, 12, 31))
    actual_active = False if is_vacant else determine_status(actual_issue_date)

    record.sbn_no = format_sbn_with_year(raw_next, actual_issue_date, is_vacant)
    record.plate_no = sanitize_plate(record.plate_no, record.chassis_no, record.motor_no)
    
    if not record.driving_route or str(record.driving_route).strip() == "":
        record.driving_route = get_dominant_driving_route(db, record.route) or record.route

    update_data = record.dict(exclude={'issue_date', 'valid_until'})
    new_record = FranchiseRecord(
        **update_data, processed_by=full_name, issue_date=actual_issue_date,
        valid_until=actual_valid_until, is_active=actual_active
    )
    db.add(new_record)
    db.commit()
    db.refresh(new_record)
    invalidate_ml_cache(new_record.route)
    
    log_action(db, full_name, "CREATE_RECORD", new_record.id, new_record.route, f"Registered new operator {record.operator_name or 'VACANT'} under SBN {record.sbn_no} ({record.route})")
    return {"status": "success", "message": f"Franchise operator added successfully with SBN {record.sbn_no}."}

@app.put("/franchise/{record_id}")
def update_franchise(record_id: str, record: FranchiseCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    db_record = db.query(FranchiseRecord).filter(FranchiseRecord.id == record_id, FranchiseRecord.is_deleted == False).first()
    if not db_record: raise HTTPException(status_code=404)
    
    raw_old_sbn = str(db_record.sbn_no).strip().upper()
    raw_new_sbn = str(record.sbn_no).strip().upper()
    
    new_base_sbn = normalize_base_sbn(record.sbn_no)
    is_renewal = False
    is_change_motor = False
    record.plate_no = sanitize_plate(record.plate_no, record.chassis_no, record.motor_no)
    
    is_vacant = not record.operator_name or str(record.operator_name).strip() == ""
    
    current_time = get_pht_now()
    manual_issue = parse_safe_date(record.issue_date, current_time.year) if record.issue_date else None
    manual_valid = parse_safe_date(record.valid_until, current_time.year) if record.valid_until else None

    if is_vacant:
        db_record.is_active = False
        db_record.issue_date = None
        db_record.valid_until = None
    else:
        if (
            str(db_record.motor_no).strip().upper() != str(record.motor_no).strip().upper() or
            str(db_record.chassis_no).strip().upper() != str(record.chassis_no).strip().upper() or
            str(db_record.make).strip().upper() != str(record.make).strip().upper()
        ):
            is_change_motor = True
            db_record.issue_date = manual_issue if manual_issue else current_time
            db_record.valid_until = manual_valid if manual_valid else datetime(db_record.issue_date.year, 12, 31)
            db_record.is_active = True
            
        elif raw_old_sbn != raw_new_sbn: 
            is_renewal = True
            db_record.issue_date = manual_issue if manual_issue else current_time
            db_record.valid_until = manual_valid if manual_valid else datetime(db_record.issue_date.year, 12, 31)
            db_record.is_active = True

        elif manual_issue: 
            db_record.issue_date = manual_issue
            db_record.valid_until = manual_valid if manual_valid else datetime(manual_issue.year, 12, 31)
            db_record.is_active = determine_status(db_record.issue_date)

    record.sbn_no = format_sbn_with_year(new_base_sbn, db_record.issue_date, is_vacant)

    if not record.driving_route or str(record.driving_route).strip() == "":
        record.driving_route = get_dominant_driving_route(db, db_record.route) or db_record.route

    update_data = record.dict(exclude={'issue_date', 'valid_until'})
    for key, value in update_data.items():
        setattr(db_record, key, value)

    db.commit()
    invalidate_ml_cache(db_record.route)
    full_name = get_full_name(current_user)
    
    if is_change_motor:
        log_action(db, full_name, "CHANGE_MOTOR", db_record.id, db_record.route, f"Processed Change Motor for {db_record.operator_name or 'VACANT'}. Updated Date Issued to {db_record.issue_date.strftime('%Y-%m-%d')}.")
    elif is_renewal:
        log_action(db, full_name, "RENEWAL", db_record.id, db_record.route, f"Renewed SBN to {new_base_sbn}. Extended to Dec 31, {get_pht_now().year}")
    else:
        log_action(db, full_name, "EDIT_RECORD", db_record.id, db_record.route, f"Updated record for {db_record.operator_name or 'VACANT'} - SBN {db_record.sbn_no} ({db_record.route})")
    return {"status": "success"}

@app.post("/upload/database")
async def upload_database_file(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    temp_db_path = os.path.join(BASE_DIR, f"temp_import_{uuid.uuid4().hex}.db")
    with open(temp_db_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    try:
        temp_engine = create_engine(f"sqlite:///{temp_db_path}")
        inspector = inspect(temp_engine)
        if "franchise_records" not in inspector.get_table_names():
            raise Exception("Uploaded database is invalid or corrupted. Missing 'franchise_records' table.")
            
        TempSession = sessionmaker(bind=temp_engine)
        temp_db = TempSession()
        imported_records = temp_db.query(FranchiseRecord).all()
        db = SessionLocal()
        
        existing_ids = {r.id for r in db.query(FranchiseRecord.id).all()}
        
        new_count = 0
        skipped_count = 0
        
        for r in imported_records:
            if getattr(r, 'is_deleted', False):
                skipped_count += 1
                continue
                
            if r.id in existing_ids:
                skipped_count += 1
                continue

            is_vac = not r.operator_name or str(r.operator_name).strip() == ""
            new_record = FranchiseRecord(
                id=r.id, sbn_no=format_sbn_with_year(r.sbn_no, r.issue_date, is_vac), operator_name=r.operator_name,
                address=r.address, motor_no=r.motor_no, chassis_no=r.chassis_no,
                make=r.make, plate_no=r.plate_no, route=r.route,
                driving_route=r.driving_route, issue_date=r.issue_date,
                valid_until=r.valid_until, is_active=r.is_active,
                processed_by=get_full_name(current_user), 
                updated_at=r.updated_at
            )
            db.add(new_record)
            existing_ids.add(r.id)
            new_count += 1
            
        db.commit()
        invalidate_ml_cache()
        temp_db.close()
        os.remove(temp_db_path)
        log_action(db, get_full_name(current_user), "DATABASE_MIGRATION", "0", "ALL", f"Merged {new_count} records from .db file. Skipped {skipped_count}.")
        return {"imported": new_count, "skipped": skipped_count}
    except Exception as e:
        if os.path.exists(temp_db_path): os.remove(temp_db_path)
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/upload/bulk/{route_name}")
async def upload_bulk_files(route_name: str, files: List[UploadFile] = File(...), current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    full_name = get_full_name(current_user)
    imported_count = 0
    current_time = get_pht_now()
    error_log = [] 
    
    existing_route_records = db.query(FranchiseRecord).filter(FranchiseRecord.route == route_name.upper(), FranchiseRecord.is_deleted == False).all()

    sbn_map = {normalize_base_sbn(pr.sbn_no): pr for pr in existing_route_records}
    plate_map = {pr.plate_no: pr for pr in existing_route_records if pr.plate_no}
    motor_map = {pr.motor_no: pr for pr in existing_route_records if pr.motor_no}
    chassis_map = {pr.chassis_no: pr for pr in existing_route_records if pr.chassis_no}

    for file in files:
        if file.filename.startswith("~$") or file.filename.startswith("._") or file.filename == ".DS_Store":
            continue

        contents = await file.read()
        try:
            if file.filename.endswith(".xlsx") or file.filename.endswith(".csv"):
                row_dicts = []
                if file.filename.endswith(".xlsx"):
                    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
                    sheet = wb.active
                    header_idx = -1
                    headers = []
                    for r_idx in range(1, min(20, sheet.max_row + 1)):
                        row_vals = [str(sheet.cell(row=r_idx, column=c).value).strip().upper() for c in range(1, sheet.max_column + 1) if sheet.cell(row=r_idx, column=c).value is not None]
                        if any("NAME" in val for val in row_vals) and any("SBN" in val for val in row_vals):
                            header_idx = r_idx
                            headers = [str(sheet.cell(row=r_idx, column=c).value).strip().upper() if sheet.cell(row=r_idx, column=c).value is not None else f"COL_{c}" for c in range(1, sheet.max_column + 1)]
                            break
                    if header_idx != -1:
                        for r_idx in range(header_idx + 1, sheet.max_row + 1):
                            row_dict = {}
                            for c_idx in range(1, sheet.max_column + 1):
                                row_dict[headers[c_idx - 1]] = sheet.cell(row=r_idx, column=c_idx).value
                            row_dicts.append(row_dict)
                else:
                    df = pd.read_csv(io.BytesIO(contents), header=None)
                    df.fillna("", inplace=True)
                    header_idx = -1
                    for i, row in df.iterrows():
                        row_vals = [str(x).upper().strip() for x in row.values if str(x).strip()]
                        if any("NAME" in val for val in row_vals) and any("SBN" in val for val in row_vals):
                            header_idx = i
                            break
                    if header_idx != -1:
                        df.columns = [str(c).strip().upper() for c in df.iloc[header_idx]]
                        df = df.iloc[header_idx+1:].reset_index(drop=True)
                        row_dicts = df.to_dict(orient="records")

                for row_dict in row_dicts:
                    raw_sbn = get_fuzzy_col_dict(row_dict, "SBN")
                    if not raw_sbn: continue 
                        
                    name = get_fuzzy_col_dict(row_dict, "NAME")
                    chassis = get_fuzzy_col_dict(row_dict, "CHASSIS")
                    motor = get_fuzzy_col_dict(row_dict, "MOTOR")
                    plate = get_fuzzy_col_dict(row_dict, "PLATE")
                    address = get_fuzzy_col_dict(row_dict, "ADDRESS")
                    make = get_fuzzy_col_dict(row_dict, "MAKE")

                    is_vacant = not name or str(name).strip() == ""

                    raw_date = None
                    if not is_vacant:
                        for col, val in row_dict.items():
                            if "RENEWAL" in str(col).upper() or "DATE" in str(col).upper():
                                raw_date = val
                                break

                    parsed_date = parse_safe_date(raw_date, current_time.year)
                    
                    incoming_base_sbn = normalize_base_sbn(raw_sbn)
                    clean_plate = sanitize_plate(plate, chassis, motor)
                    
                    existing_record = sbn_map.get(incoming_base_sbn)

                    if existing_record:
                        if get_record_year(parsed_date) >= get_record_year(existing_record.issue_date):
                            existing_record.issue_date = parsed_date
                            existing_record.valid_until = datetime(parsed_date.year, 12, 31) if parsed_date else None
                            existing_record.is_active = False if is_vacant else determine_status(parsed_date)
                            existing_record.sbn_no = format_sbn_with_year(incoming_base_sbn, parsed_date, is_vacant)
                            
                            if name and str(name).strip():
                                existing_record.operator_name = str(name).strip().upper()
                            existing_record.address = set_if_blank(existing_record.address, address)
                            existing_record.make = set_if_blank(existing_record.make, make)
                            existing_record.plate_no = set_if_blank(existing_record.plate_no, clean_plate)
                            existing_record.chassis_no = set_if_blank(existing_record.chassis_no, chassis)
                            existing_record.motor_no = set_if_blank(existing_record.motor_no, motor)
                            imported_count += 1
                    else:
                        record = FranchiseRecord(
                            sbn_no=format_sbn_with_year(incoming_base_sbn, parsed_date, is_vacant),
                            operator_name=name.upper() if name else "",
                            address=address.upper() if address else "", motor_no=motor.upper() if motor else "",
                            plate_no=clean_plate.upper() if clean_plate else "", chassis_no=chassis.upper() if chassis else "",
                            make=make.upper() if make else "", route=route_name.upper(), driving_route="", 
                            issue_date=parsed_date, valid_until=datetime(parsed_date.year, 12, 31) if parsed_date else None,
                            processed_by=full_name, is_active=False if is_vacant else determine_status(parsed_date)
                        )
                        db.add(record)
                        sbn_map[incoming_base_sbn] = record
                        imported_count += 1

            elif file.filename.endswith(".docx"):
                extracted = extract_docx_data(contents, route_name.upper(), current_time.year)
                is_vacant = not extracted['operator_name'] or str(extracted['operator_name']).strip() == ""
                issue_date = parse_safe_date(extracted['issue_date'], current_time.year)
                
                clean_plate = sanitize_plate(extracted['plate_no'], extracted['chassis_no'], extracted['motor_no'])
                incoming_base_sbn = normalize_base_sbn(extracted['sbn_no'])
                
                existing_record = sbn_map.get(incoming_base_sbn)

                if not existing_record:
                    if clean_plate and str(clean_plate).strip(): existing_record = plate_map.get(str(clean_plate).strip().upper())
                    if not existing_record and extracted['motor_no']: existing_record = motor_map.get(str(extracted['motor_no']).strip().upper())
                    if not existing_record and extracted['chassis_no']: existing_record = chassis_map.get(str(extracted['chassis_no']).strip().upper())
                    
                    if existing_record:
                        incoming_base_sbn = normalize_base_sbn(existing_record.sbn_no)

                if existing_record:
                    if extracted['operator_name'] and str(extracted['operator_name']).strip():
                        existing_record.operator_name = str(extracted['operator_name']).strip().upper()
                    if extracted['address'] and str(extracted['address']).strip():
                        existing_record.address = str(extracted['address']).strip().upper()
                    if extracted['make'] and str(extracted['make']).strip():
                        existing_record.make = str(extracted['make']).strip().upper()
                    if extracted.get('driving_route') and str(extracted['driving_route']).strip():
                        existing_record.driving_route = str(extracted['driving_route']).strip().upper()
                    if clean_plate and str(clean_plate).strip():
                        existing_record.plate_no = str(clean_plate).strip().upper()
                    if extracted['chassis_no'] and str(extracted['chassis_no']).strip():
                        existing_record.chassis_no = str(extracted['chassis_no']).strip().upper()
                    if extracted['motor_no'] and str(extracted['motor_no']).strip():
                        existing_record.motor_no = str(extracted['motor_no']).strip().upper()

                    doc_year = get_record_year(issue_date)
                    existing_year = get_record_year(existing_record.issue_date)
                    if issue_date and (not existing_record.issue_date or doc_year > existing_year):
                        existing_record.issue_date = issue_date
                        existing_record.valid_until = datetime(issue_date.year, 12, 31)
                        existing_record.is_active = False if is_vacant else determine_status(issue_date)
                    
                    existing_record.sbn_no = format_sbn_with_year(incoming_base_sbn, existing_record.issue_date, is_vacant)
                    imported_count += 1
                else:
                    # FIX: Safely extract integer to drop true placeholders instead of blindly rejecting if it contains "000"
                    sbn_int = extract_sbn_integer(incoming_base_sbn)
                    if sbn_int is None or sbn_int != 0:
                        record = FranchiseRecord(
                            sbn_no=format_sbn_with_year(incoming_base_sbn, issue_date, is_vacant),
                            operator_name=extracted['operator_name'].upper() if extracted['operator_name'] else "",
                            address=extracted['address'].upper() if extracted['address'] else "",
                            motor_no=extracted['motor_no'].upper() if extracted['motor_no'] else "",
                            plate_no=clean_plate.upper() if clean_plate else "",
                            chassis_no=extracted['chassis_no'].upper() if extracted['chassis_no'] else "",
                            make=extracted['make'].upper() if extracted['make'] else "",
                            route=route_name.upper(),
                            driving_route=extracted['driving_route'].upper() if extracted.get('driving_route') else "",
                            issue_date=issue_date,
                            valid_until=datetime(issue_date.year, 12, 31) if issue_date else None,
                            processed_by=full_name,
                            is_active=False if is_vacant else determine_status(issue_date)
                        )
                        db.add(record)
                        sbn_map[incoming_base_sbn] = record
                        imported_count += 1
        except Exception as e:
            error_msg = f"Failed extracting {file.filename}: {str(e)}"
            force_log(error_msg)
            error_log.append(error_msg) 
            continue

    db.commit()
    invalidate_ml_cache(route_name)

    dominant_route = get_dominant_driving_route(db, route_name)
    if dominant_route:
        for pr in db.query(FranchiseRecord).filter(FranchiseRecord.route == route_name.upper(), FranchiseRecord.is_deleted == False).all():
            curr_dr = str(pr.driving_route).strip().upper() if pr.driving_route else ""
            if curr_dr == "" or (curr_dr == "POBLACION" and dominant_route != "POBLACION"):
                pr.driving_route = dominant_route
        db.commit()

    log_action(db, "SYSTEM_MIGRATION", "IMPORT", "0", route_name.upper(), f"Imported/Updated {imported_count} records. Errors: {len(error_log)}")
    
    return {
        "imported": imported_count,
        "errors": error_log
    }

@app.post("/franchise/download/word/{record_id}")
async def download_word_doc(record_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    record = db.query(FranchiseRecord).filter(FranchiseRecord.id == record_id).first()
    if not record: raise HTTPException(status_code=404)
    settings = init_settings(db)
    
    is_vacant = not record.operator_name or str(record.operator_name).strip() == ""
    full_sbn = format_sbn_with_year(record.sbn_no, record.issue_date, is_vacant)
    
    cert_data = {
        "sbn_no": full_sbn, "operator_name": record.operator_name,
        "address": record.address, "motor_no": record.motor_no,
        "chassis_no": record.chassis_no, "make": record.make,
        "plate_no": record.plate_no, "route": record.route,
        "driving_route": record.driving_route,
        "issue_date": record.issue_date, "valid_until": record.valid_until
    }
    committee_data = {"committee_chair": settings.committee_chair}
    
    doc_path, media_type = await asyncio.to_thread(generate_certificate, cert_data, committee_data, return_format="docx")
    
    if os.path.exists(doc_path):
        log_action(db, get_full_name(current_user), "DOWNLOAD_DOCX", record.id, record.route, f"Downloaded raw MTOP Word Document for SBN {full_sbn}")
        headers = {'Content-Disposition': f'attachment; filename="{full_sbn}.docx"'}
        return FileResponse(path=doc_path, headers=headers, media_type=media_type)
        
    raise HTTPException(status_code=500)

@app.post("/franchise/generate/{record_id}")
async def generate_doc(record_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    record = db.query(FranchiseRecord).filter(FranchiseRecord.id == record_id).first()
    if not record: raise HTTPException(status_code=404)
    settings = init_settings(db)
    
    is_vacant = not record.operator_name or str(record.operator_name).strip() == ""
    full_sbn = format_sbn_with_year(record.sbn_no, record.issue_date, is_vacant)
    
    cert_data = {
        "sbn_no": full_sbn, "operator_name": record.operator_name,
        "address": record.address, "motor_no": record.motor_no,
        "chassis_no": record.chassis_no, "make": record.make,
        "plate_no": record.plate_no, "route": record.route,
        "driving_route": record.driving_route,
        "issue_date": record.issue_date, "valid_until": record.valid_until
    }
    committee_data = {"committee_chair": settings.committee_chair}

    doc_path, media_type = await asyncio.to_thread(generate_certificate, cert_data, committee_data, return_format="pdf")
    
    if os.path.exists(doc_path):
        log_action(db, get_full_name(current_user), "PRINT_MTOP", record.id, record.route, f"Generated MTOP Document for SBN {full_sbn}")
        headers = {'Content-Disposition': f'attachment; filename="{full_sbn}.pdf"'}
        return FileResponse(path=doc_path, headers=headers, media_type=media_type)
        
    raise HTTPException(status_code=500)

@app.get("/logs/record/{record_id}")
def get_record_history(record_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    logs = db.query(AuditLog).filter(AuditLog.target_id == record_id).order_by(AuditLog.timestamp.desc()).all()
    result = []
    
    for log in logs:
        log_dict = {c.name: getattr(log, c.name) for c in log.__table__.columns}
        
        cname = str(log_dict.get("clerk_name", "")).strip()
        if not cname or cname.lower() in ["none", "null"]:
            cname = "SYSTEM ADMIN"
            
        log_dict["clerk_name"] = cname
        log_dict["user"] = cname
        
        result.append(log_dict)
        
    return result

@app.get("/logs")
def get_audit_logs(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    logs = db.query(AuditLog).order_by(AuditLog.timestamp.desc()).limit(200).all()
    result = []
    
    for log in logs:
        log_dict = {c.name: getattr(log, c.name) for c in log.__table__.columns}
        
        tid = str(log_dict.get("target_id", ""))
        if len(tid) > 10:  
            rec = db.query(FranchiseRecord).filter(FranchiseRecord.id == tid).first()
            if rec:
                val = f"{rec.sbn_no} ({rec.operator_name or 'VACANT'})"
            else:
                val = "Unknown / Deleted Record"
        else:
            val = tid if tid and tid != "0" else "SYSTEM"
            
        log_dict["target_id"] = val
        log_dict["target_record"] = val 
        
        cname = str(log_dict.get("clerk_name", "")).strip()
        if not cname or cname.lower() in ["none", "null"]:
            cname = "SYSTEM ADMIN"
            
        log_dict["clerk_name"] = cname
        log_dict["user"] = cname 
        
        result.append(log_dict)
        
    return result

@app.get("/franchise/route/{route_name}")
def get_route_records(route_name: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    route_upper = route_name.upper()
    
    # CASE 1: Fetch ALL Routes
    if route_upper == "ALL":
        records = db.query(FranchiseRecord).filter(FranchiseRecord.is_deleted == False).all()
        unique_map = {}
        for r in records:
            base_id = normalize_base_sbn(r.sbn_no)
            r.sbn_no = base_id
            route_key = (r.route, base_id)
            if route_key not in unique_map:
                unique_map[route_key] = r
            else:
                curr = unique_map[route_key]
                curr_year = get_record_year(curr.issue_date)
                new_year = get_record_year(r.issue_date)
                if new_year > curr_year:
                    unique_map[route_key] = r
                elif new_year == curr_year and (r.operator_name and str(r.operator_name).strip() and not str(curr.operator_name).strip()):
                    unique_map[route_key] = r
        deduped_records = list(unique_map.values())
        
    # CASE 2: Fetch CUSTOM routes (Comma separated)
    elif "," in route_upper:
        route_list = [r.strip() for r in route_upper.split(",")]
        records = db.query(FranchiseRecord).filter(
            FranchiseRecord.route.in_(route_list), 
            FranchiseRecord.is_deleted == False
        ).all()
        unique_map = {}
        for r in records:
            base_id = normalize_base_sbn(r.sbn_no)
            r.sbn_no = base_id
            route_key = (r.route, base_id)
            if route_key not in unique_map:
                unique_map[route_key] = r
            else:
                curr = unique_map[route_key]
                curr_year = get_record_year(curr.issue_date)
                new_year = get_record_year(r.issue_date)
                if new_year > curr_year:
                    unique_map[route_key] = r
                elif new_year == curr_year and (r.operator_name and str(r.operator_name).strip() and not str(curr.operator_name).strip()):
                    unique_map[route_key] = r
        deduped_records = list(unique_map.values())
        
    # CASE 3: Fetch SINGLE Exact Route (Current behavior)
    else:
        records = db.query(FranchiseRecord).filter(
            FranchiseRecord.route == route_upper, 
            FranchiseRecord.is_deleted == False
        ).all()
        deduped_records = deduplicate_records_by_base_sbn(records)

    # Format the SBNs and attach status
    for record in deduped_records:
        is_vac = not record.operator_name or str(record.operator_name).strip() == ""
        record.sbn_no = format_sbn_with_year(record.sbn_no, record.issue_date, is_vacant=is_vac)

    sorted_records = sorted(deduped_records, key=get_sbn_sort_key)
    return [attach_status(r) for r in sorted_records]

@app.get("/franchise/status/inactive")
def get_inactive_records(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    records = db.query(FranchiseRecord).filter(FranchiseRecord.is_active == False, FranchiseRecord.is_deleted == False).all()
    deduped = deduplicate_records_by_base_sbn(records)
    for r in deduped:
        is_vac = not r.operator_name or str(r.operator_name).strip() == ""
        r.sbn_no = format_sbn_with_year(r.sbn_no, r.issue_date, is_vac)
    return [attach_status(r) for r in sorted(deduped, key=get_sbn_sort_key)]

@app.get("/predict/{route}")
def get_prediction(route: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return run_kmeans_clustering(db, route)

@app.get("/export/masterlist/{route_name}")
def export_toda_masterlist(route_name: str, status_filter: str = "ALL", current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if route_name.upper() == "ALL":
        records_query = db.query(FranchiseRecord).filter(FranchiseRecord.is_deleted == False).all()
    else:
        records_query = db.query(FranchiseRecord).filter(FranchiseRecord.route == route_name.upper(), FranchiseRecord.is_deleted == False).all()
        
    records_query = deduplicate_records_by_base_sbn(records_query)
    records_query = sorted(records_query, key=get_sbn_sort_key)
    current_year = get_pht_now().year
    
    filtered_records = []
    for r in records_query:
        computed_filter_status = compute_record_status(r)
        if status_filter == "ALL" or status_filter == computed_filter_status:
            filtered_records.append(r)

    if not filtered_records:
        raise HTTPException(status_code=404, detail="No records found matching this filter")
        
    csv_data = []
    for r in filtered_records:
        is_vacant = not r.operator_name or str(r.operator_name).strip() == ""
        formatted_sbn = normalize_base_sbn(r.sbn_no)
        
        csv_data.append({
            "SBN NO.": escape_excel(formatted_sbn),
            "DATE OF RENEWAL": None if is_vacant else r.issue_date,
            "NAME": "" if is_vacant else escape_excel(r.operator_name),
            "MAKE": escape_excel(r.make), 
            "MOTOR NO.": escape_excel(r.motor_no),
            "CHASSIS NO.": escape_excel(r.chassis_no),
            "PLATE NO.": escape_excel(r.plate_no)
        })
        
    df = pd.DataFrame(csv_data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Masterlist", startrow=1)
        ws = writer.sheets['Masterlist']
        
        calibri_11 = Font(name="Calibri", size=11)
        calibri_11_bold = Font(name="Calibri", size=11, bold=True)
        calibri_16_bold = Font(name="Calibri", size=16, bold=True)
        
        fill_vacant = PatternFill(start_color="FF00B0F0", end_color="FF00B0F0", fill_type="solid")
        fill_y2_1 = PatternFill(start_color="FFBDD7EE", end_color="FFBDD7EE", fill_type="solid")
        fill_y3_7 = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
        fill_y8 = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
        
        for row in ws.iter_rows():
            for cell in row:
                cell.font = calibri_11

        total_row_count = len(filtered_records) + 2

        ws.merge_cells('A1:B1')
        ws['A1'] = f'=COUNTIFS(B3:B{total_row_count}, ">="&K5, B3:B{total_row_count}, "<="&K6)'
        ws['A1'].font = calibri_16_bold
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        for col_num in range(1, 8):
            cell = ws.cell(row=2, column=col_num)
            cell.font = calibri_11_bold
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws['I3'] = "VACANT"
        ws['I3'].font = calibri_11_bold
        ws['I3'].fill = fill_vacant
        
        ws['I4'] = "2024-2025"
        ws['I4'].font = calibri_11_bold
        ws['I4'].fill = fill_y2_1
        
        ws['I5'] = "2019-2023"
        ws['I5'].font = calibri_11_bold
        ws['I5'].fill = fill_y3_7
        
        ws['I6'] = "2018"
        ws['I6'].font = calibri_11_bold
        ws['I6'].fill = fill_y8

        ws['J5'] = "Count:"
        ws['J5'].font = calibri_11_bold
        ws['K5'] = f"{current_year}-01-01"
        ws['K5'].font = calibri_11
        ws['K6'] = f"{current_year}-12-31"
        ws['K6'].font = calibri_11

        for row_num, r_dict in enumerate(csv_data, start=3):
            op_name = str(r_dict["NAME"]).strip()
            issue_date = r_dict["DATE OF RENEWAL"]
            issue_year = issue_date.year if pd.notna(issue_date) else 0
            
            fill_color = None
            if not op_name or op_name == "'":
                fill_color = fill_vacant
            elif issue_year == current_year:
                fill_color = None
            elif issue_year == current_year - 1 or issue_year == current_year - 2:
                fill_color = fill_y2_1
            elif current_year - 7 <= issue_year <= current_year - 3:
                fill_color = fill_y3_7
            else:
                fill_color = fill_y8
            
            for col_num in range(1, 8):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = calibri_11
                if fill_color:
                    cell.fill = fill_color
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col_num == 2 and cell.value:
                    cell.number_format = 'dd-mmm-yy'
        
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in ['I', 'J', 'K']:
                ws.column_dimensions[col_letter].width = 16
                continue
            
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if col_idx == 2 and cell_value: 
                    if 12 > max_length: max_length = 12
                elif cell_value:
                    length = len(str(cell_value))
                    if length > max_length:
                        max_length = length
            ws.column_dimensions[col_letter].width = max_length + 2

    output.seek(0)
    
    if status_filter == "ALL":
        filename_str = f"{route_name.upper()} {current_year}.xlsx"
    else:
        filename_str = f"{route_name.upper()} {current_year} - {status_filter.upper()}.xlsx"

    log_action(db, get_full_name(current_user), "EXPORT_MASTERLIST", "0", route_name.upper(), f"Exported {status_filter} Masterlist for {route_name.upper()}")
    
    headers = {
        'Content-Disposition': f'attachment; filename="{filename_str}"'
    }
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@app.get("/api/route-distribution")
def get_route_distribution(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    current_year = get_pht_now().year
    deduped_all = get_all_deduplicated_records(db)
    route_counts = {}
    for r in deduped_all:
        if r.issue_date and r.issue_date.year == current_year and r.operator_name and str(r.operator_name).strip() != "":
            route_counts[r.route] = route_counts.get(r.route, 0) + 1
    route_data = [{"route": route, "count": count} for route, count in sorted(route_counts.items())]
    return route_data

@app.get("/api/stats")
def get_api_stats(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return get_global_stats(db)

@app.get("/api/dashboard")
def get_api_dashboard(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return get_global_stats(db)

@app.get("/stats/global")
def get_global_stats(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    current_time = get_pht_now()
    current_year = current_time.year
    current_month = current_time.month
    today_str = current_time.strftime('%Y-%m-%d')
    start_of_week = (current_time - timedelta(days=current_time.weekday())).strftime('%Y-%m-%d')
    
    deduped_all = get_all_deduplicated_records(db)
    total_system_capacity = len(deduped_all)
    vacant_slots = sum(1 for r in deduped_all if not r.operator_name or str(r.operator_name).strip() == "")
    
    active_operators = sum(1 for r in deduped_all if r.issue_date and r.issue_date.year == current_year and r.operator_name and str(r.operator_name).strip() != "")
    flagged_pending = sum(1 for r in deduped_all if r.issue_date and r.issue_date.year == current_year - 1 and r.operator_name and str(r.operator_name).strip() != "")
    inactive_and_vacant = sum(1 for r in deduped_all if (not r.operator_name or str(r.operator_name).strip() == "") or (r.issue_date and r.issue_date.year <= current_year - 2) or not r.issue_date)

    daily_apps = sum(1 for r in deduped_all if r.issue_date and r.issue_date.strftime('%Y-%m-%d') == today_str)
    weekly_apps = sum(1 for r in deduped_all if r.issue_date and r.issue_date.strftime('%Y-%m-%d') >= start_of_week)
    monthly_apps = sum(1 for r in deduped_all if r.issue_date and r.issue_date.year == current_year and r.issue_date.month == current_month)
    yearly_apps = active_operators

    route_counts = {}
    for r in deduped_all:
        if r.issue_date and r.issue_date.year == current_year and r.operator_name and str(r.operator_name).strip() != "":
            route_counts[r.route] = route_counts.get(r.route, 0) + 1
    route_data = [{"route": route, "count": count} for route, count in sorted(route_counts.items())]

    days_map = {'0': 'Sun', '1': 'Mon', '2': 'Tue', '3': 'Wed', '4': 'Thu', '5': 'Fri', '6': 'Sat'}
    daily_trend_dict = {d: 0 for d in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']}
    
    for r in deduped_all:
        if r.issue_date:
            day_idx = r.issue_date.strftime('%w')
            if day_idx in days_map:
                daily_trend_dict[days_map[day_idx]] += 1
            
    daily_trend = [{"name": k, "val": v} for k, v in daily_trend_dict.items()]
        
    weekly_trend = []
    for i in range(4, -1, -1):
        target_date = current_time - timedelta(days=current_time.weekday() + (i * 7))
        start_str = target_date.strftime('%Y-%m-%d')
        end_str = (target_date + timedelta(days=6)).strftime('%Y-%m-%d')
        count = sum(1 for r in deduped_all if r.issue_date and start_str <= r.issue_date.strftime('%Y-%m-%d') <= end_str)
        weekly_trend.append({"name": target_date.strftime('%b %d'), "val": count})

    monthly_trend = []
    for i in range(5, -1, -1):
        target_month = current_month - i
        target_year = current_year
        if target_month <= 0:
            target_month += 12
            target_year -= 1
        month_name = calendar.month_abbr[target_month]
        count = sum(1 for r in deduped_all if r.issue_date and r.issue_date.year == target_year and r.issue_date.month == target_month)
        monthly_trend.append({"name": month_name, "val": count})

    return {
        "total_system_capacity": total_system_capacity,
        "vacant_slots": inactive_and_vacant,
        "daily_apps": daily_apps, "weekly_apps": weekly_apps, "monthly_apps": monthly_apps, "yearly_apps": yearly_apps,
        "flagged_pending": flagged_pending, "revoked": inactive_and_vacant, 
        "route_breakdown": route_data,
        "daily_trend": daily_trend,
        "weekly_trend": weekly_trend,
        "monthly_trend": monthly_trend
    }

def kill_zombie_port(port):
    try:
        if platform.system() == "Windows":
            # FIXED FOR WINDOWS - Avoid OSError by passing DEVNULL to handles
            output = subprocess.check_output(f"netstat -ano | findstr :{port}", shell=True, stdin=subprocess.DEVNULL, stderr=subprocess.DEVNULL).decode()
            for line in output.strip().split('\n'):
                if "LISTENING" in line:
                    pid = line.strip().split()[-1]
                    if str(pid) != str(os.getpid()):
                        force_log(f"CRITICAL: Found ghost process (PID: {pid}). Nuking it...")
                        subprocess.run(f"taskkill /PID {pid} /F", shell=True, stdin=subprocess.DEVNULL, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        else:
            try:
                output = subprocess.check_output(f"lsof -t -i:{port}", shell=True).decode()
                for pid in output.strip().split('\n'):
                    if pid and str(pid) != str(os.getpid()):
                        force_log(f"CRITICAL: Found ghost process (PID: {pid}). Nuking it...")
                        os.system(f"kill -9 {pid}")
            except Exception:
                output = subprocess.check_output(f"fuser {port}/tcp 2>/dev/null", shell=True).decode()
                for pid in output.strip().split():
                    if pid and str(pid) != str(os.getpid()):
                        force_log(f"CRITICAL: Found ghost process (PID: {pid}). Nuking it...")
                        os.system(f"kill -9 {pid}")
    except Exception as e:
        pass

if __name__ == "__main__":
    try:
        force_log("Scanning for zombie processes on Port 43888...")
        kill_zombie_port(43888)
        force_log("Port 43888 is clear. Starting Uvicorn server...")
        print("[Backend] Starting Uvicorn server on http://0.0.0.0:43888 ...")
        sys.stdout.flush()
        uvicorn.run(app, host="0.0.0.0", port=43888)
    except Exception as e:
        err_str = traceback.format_exc()
        force_log(f"\n[{datetime.now()}] !!! SERVER EXECUTION FATAL ERROR !!!\n{err_str}")
        print(f"[Backend Fatal Error]: {err_str}", file=sys.stderr)
        sys.stderr.flush()
        sys.exit(1)